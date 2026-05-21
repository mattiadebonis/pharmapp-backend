#!/usr/bin/env python3
"""Import AIFA "liste di trasparenza" (classi di rimborsabilità A/H/C).

Le liste sono pubblicate mensilmente da AIFA in formato Excel sul
portale https://www.aifa.gov.it/liste-farmaci-a-h e includono:
  - CODICE_AIC  (9 cifre, chiave package)
  - DENOMINAZIONE / DESCRIZIONE
  - CLASSE       (A / A-PHT / H / C / C-bis / Cnn)
  - PREZZO       (prezzo di riferimento EUR)
  - NOTA AIFA    (es. "Nota 13")

Lo script:
  1. Legge il file CSV o XLSX
  2. Normalizza le colonne (le intestazioni variano leggermente
     mese su mese — gestiamo i sinonimi più comuni)
  3. Upsert su `catalog_it_packages.reimbursement_class`,
     `reference_price`, `reimbursement_note`,
     `reimbursement_updated_at`

Usage:
    # CSV con header italiano
    python scripts/import_aifa_trasparenza.py --file liste-A.csv

    # XLSX (richiede openpyxl)
    python scripts/import_aifa_trasparenza.py --file liste-A.xlsx --sheet 0

    # Dry run (parsing only, no DB writes)
    python scripts/import_aifa_trasparenza.py --file liste-A.csv --dry-run

    # Solo classe (ignora prezzi)
    python scripts/import_aifa_trasparenza.py --file liste-A.csv --class-only

Where to download:
    https://www.aifa.gov.it/liste-farmaci-a-h
    Pagina "Lista trasparenza farmaci equivalenti" — generalmente
    aggiornata il 15 di ogni mese.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))


# ─── Classe AIFA: normalizzazione ─────────────────────────────────────────────

# Lo schema (migration 034) accetta: A, A-PHT, H, C, C-bis, Cnn.
# Nelle liste AIFA il valore è scritto in modi vari ("A", "Classe A",
# "A-NN", "PHT", ecc.). Normalizziamo qui.
_CLASS_NORMALIZATION = [
    (re.compile(r"^A[\s\-]?PHT$", re.IGNORECASE), "A-PHT"),
    (re.compile(r"^PHT$", re.IGNORECASE), "A-PHT"),
    (re.compile(r"^A[\s\-]?NN$", re.IGNORECASE), "Cnn"),  # alcuni list label
    (re.compile(r"^Cnn$", re.IGNORECASE), "Cnn"),
    (re.compile(r"^C[\s\-]?bis$", re.IGNORECASE), "C-bis"),
    (re.compile(r"^Classe\s+A$", re.IGNORECASE), "A"),
    (re.compile(r"^Classe\s+H$", re.IGNORECASE), "H"),
    (re.compile(r"^Classe\s+C$", re.IGNORECASE), "C"),
    (re.compile(r"^A$", re.IGNORECASE), "A"),
    (re.compile(r"^H$", re.IGNORECASE), "H"),
    (re.compile(r"^C$", re.IGNORECASE), "C"),
]


def normalize_class(value: str | None) -> str | None:
    """Normalizza la classe AIFA. Restituisce None se sconosciuta."""
    if not value:
        return None
    v = value.strip()
    if not v:
        return None
    for pattern, canonical in _CLASS_NORMALIZATION:
        if pattern.match(v):
            return canonical
    return None


# ─── Colonne: sinonimi noti ───────────────────────────────────────────────────

# Le intestazioni AIFA variano tra le pubblicazioni. Riconosciamo sia
# il formato verboso ("CODICE AIC") sia quello compatto ("AIC"), con
# o senza punto / spazio. Confronto case-insensitive su una versione
# normalizzata (lower + senza punteggiatura).
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "codice_aic": (
        "codice_aic", "aic", "codice aic", "n_aic", "n aic",
        "codice", "id confezione",
    ),
    "reimbursement_class": (
        "classe", "classe_di_rimborsabilita", "classe di rimborsabilita",
        "classe rimborso", "fascia", "fascia_di_rimborsabilita",
        "fascia di rimborsabilita", "classe rimborsabilita",
    ),
    "reference_price": (
        "prezzo", "prezzo_di_riferimento", "prezzo di riferimento",
        "prezzo riferimento", "prezzo_pubblico", "prezzo pubblico",
        "prezzo_al_pubblico", "prezzo al pubblico", "prezzo_ssn",
    ),
    "reimbursement_note": (
        "nota", "nota_aifa", "nota aifa", "note", "note_aifa",
    ),
}


def _norm_header(h: str) -> str:
    """Normalizza un header: lowercase, sostituisce non-alfanumerico
    con underscore, rimuove duplicati e bordi."""
    h = h.strip().lower()
    h = re.sub(r"[^a-z0-9]+", "_", h)
    h = h.strip("_")
    return h


def detect_columns(headers: list[str]) -> dict[str, int]:
    """Mappa nome_campo_normalizzato → indice_colonna nel CSV."""
    normalized = [_norm_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        normalized_aliases = {_norm_header(a) for a in aliases}
        for i, h in enumerate(normalized):
            if h in normalized_aliases:
                mapping[canonical] = i
                break
    return mapping


# ─── Parsing prezzo italiano ──────────────────────────────────────────────────


def parse_price(value: str | None) -> float | None:
    """Parse a price string: '12,34' or '12.34' or '1.234,56' → float."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s in {"-", "n.d.", "N.D.", "ND", "nd"}:
        return None
    # Italian "1.234,56" → "1234.56"
    if re.fullmatch(r"\d{1,3}(?:\.\d{3})+,\d{1,4}", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# ─── Stats ────────────────────────────────────────────────────────────────────


@dataclass
class Stats:
    total_rows: int = 0
    parsed: int = 0
    skipped_no_aic: int = 0
    skipped_no_class: int = 0
    skipped_bad_class: int = 0
    classes: Counter[str] = field(default_factory=Counter)

    def summary(self) -> str:
        cls_lines = "\n".join(f"    {k}: {v}" for k, v in sorted(self.classes.items()))
        return (
            f"\n{'='*60}\n"
            f"  Import Liste Trasparenza AIFA — Riepilogo\n"
            f"{'='*60}\n"
            f"  Righe lette:                 {self.total_rows}\n"
            f"  Confezioni con classe:       {self.parsed}\n"
            f"  Skippate (no AIC):           {self.skipped_no_aic}\n"
            f"  Skippate (no classe):        {self.skipped_no_class}\n"
            f"  Skippate (classe non std):   {self.skipped_bad_class}\n"
            f"  Distribuzione classi:\n{cls_lines}\n"
            f"{'='*60}"
        )


# ─── Lettura file ─────────────────────────────────────────────────────────────


def read_rows(file_path: Path, sheet: int | str = 0) -> tuple[list[str], list[list[str]]]:
    """Legge CSV o XLSX e restituisce (headers, rows)."""
    suffix = file_path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        delimiter = ";" if suffix == ".csv" else "\t"
        # AIFA usa ";" come delimiter standard
        with open(file_path, encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=delimiter, quotechar='"')
            headers = next(reader)
            rows = list(reader)
        return headers, rows

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("ERROR: openpyxl non installato. pip install openpyxl", file=sys.stderr)
            sys.exit(1)
        wb = load_workbook(file_path, data_only=True, read_only=True)
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
        all_rows = [[("" if c is None else str(c).strip()) for c in row] for row in ws.iter_rows(values_only=True)]
        if not all_rows:
            return [], []
        return all_rows[0], all_rows[1:]

    raise ValueError(f"Estensione file non supportata: {suffix}")


# ─── Parsing principale ───────────────────────────────────────────────────────


def parse_file(file_path: Path, sheet: int | str, stats: Stats,
               class_only: bool = False) -> list[dict[str, Any]]:
    """Parse il file e ritorna le righe da upsertare."""
    headers, raw_rows = read_rows(file_path, sheet=sheet)
    if not headers:
        print(f"File vuoto: {file_path}")
        return []

    columns = detect_columns(headers)
    print(f"Header rilevati: {headers}")
    print(f"Mapping colonne canoniche: {columns}")

    if "codice_aic" not in columns:
        print("ERROR: colonna AIC non trovata. Verifica header del file.", file=sys.stderr)
        print(f"Header normalizzati: {[_norm_header(h) for h in headers]}", file=sys.stderr)
        sys.exit(1)
    if "reimbursement_class" not in columns:
        print("ERROR: colonna classe non trovata. Verifica header del file.", file=sys.stderr)
        sys.exit(1)

    aic_col = columns["codice_aic"]
    class_col = columns["reimbursement_class"]
    price_col = columns.get("reference_price")
    note_col = columns.get("reimbursement_note")

    now = datetime.now(timezone.utc).isoformat()
    out: list[dict[str, Any]] = []

    for row_num, row in enumerate(raw_rows, start=2):
        stats.total_rows += 1
        if len(row) <= max(aic_col, class_col):
            continue

        codice_aic = (row[aic_col] or "").strip()
        # AIFA a volte scrive l'AIC con leading zero strippato (numero in Excel).
        # Lo schema usa TEXT a 9 cifre — riallineiamo con zfill se necessario.
        codice_aic = re.sub(r"\D", "", codice_aic)
        if codice_aic and len(codice_aic) < 9:
            codice_aic = codice_aic.zfill(9)
        if not codice_aic:
            stats.skipped_no_aic += 1
            continue

        raw_class = (row[class_col] or "").strip()
        klass = normalize_class(raw_class)
        if not raw_class:
            stats.skipped_no_class += 1
            continue
        if not klass:
            stats.skipped_bad_class += 1
            if stats.skipped_bad_class <= 5:
                print(f"  WARN row {row_num}: classe sconosciuta {raw_class!r}")
            continue

        stats.classes[klass] += 1
        stats.parsed += 1

        record: dict[str, Any] = {
            "codice_aic": codice_aic,
            "reimbursement_class": klass,
            "reimbursement_updated_at": now,
        }

        if not class_only:
            if price_col is not None and price_col < len(row):
                price = parse_price(row[price_col])
                if price is not None:
                    record["reference_price"] = price
            if note_col is not None and note_col < len(row):
                note = (row[note_col] or "").strip()
                if note:
                    record["reimbursement_note"] = note

        out.append(record)

    return out


# ─── Supabase update ──────────────────────────────────────────────────────────


def _load_env() -> tuple[str, str]:
    env_path = PROJECT_ROOT / ".env"
    env_vars: dict[str, str] = {}
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                env_vars[k.strip()] = v.strip().strip('"').strip("'")
    url = os.environ.get("SUPABASE_URL") or env_vars.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or env_vars.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY required in .env", file=sys.stderr)
        sys.exit(1)
    return url, key


def update_packages(records: list[dict[str, Any]], batch_size: int = 500) -> int:
    """Update reimbursement_* fields su catalog_it_packages by codice_aic.

    Usiamo UPDATE singolo per evitare di sovrascrivere altri campi
    della riga (un upsert con questi soli campi sostituirebbe i NOT
    NULL come cod_farmaco con errore). Per scalare: batch di
    UPDATE...IN... in transazione.
    """
    if not records:
        return 0

    url, key = _load_env()
    from supabase import ClientOptions, create_client
    supabase = create_client(url, key, options=ClientOptions(postgrest_client_timeout=300))

    updated = 0
    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        # Strategia: una RPC server-side sarebbe più efficiente. Per
        # ora aggiorniamo uno-a-uno; con 8-15k righe è accettabile
        # (~1-2 min). Documentato come TODO se si vuole accelerare.
        for rec in batch:
            aic = rec.pop("codice_aic")
            try:
                result = supabase.table("catalog_it_packages") \
                    .update(rec).eq("codice_aic", aic).execute()
                if result.data:
                    updated += 1
            except Exception as e:
                print(f"  WARN: update AIC {aic} fallito: {e!r}")
        done = min(i + batch_size, len(records))
        print(f"  trasparenza: {done}/{len(records)} (updated: {updated})")

    return updated


# ─── Main ─────────────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(description="Import AIFA liste di trasparenza (CSV/XLSX)")
    parser.add_argument("--file", type=Path, required=True,
                        help="Path al file lista trasparenza (.csv, .tsv, .xlsx)")
    parser.add_argument("--sheet", default=0,
                        help="Indice (int) o nome sheet per file XLSX (default: 0)")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
    parser.add_argument("--class-only", action="store_true",
                        help="Aggiorna solo reimbursement_class (ignora prezzo e nota)")
    parser.add_argument("--batch-size", type=int, default=500)
    args = parser.parse_args()

    if not args.file.exists():
        print(f"ERROR: File not found: {args.file}", file=sys.stderr)
        return 1

    print(f"AIFA Trasparenza Import — file: {args.file}")
    print(f"  dry_run={args.dry_run}, class_only={args.class_only}")

    stats = Stats()
    t0 = time.time()

    # Parse sheet index/name
    sheet: int | str
    try:
        sheet = int(args.sheet)
    except (ValueError, TypeError):
        sheet = args.sheet

    records = parse_file(args.file, sheet=sheet, stats=stats, class_only=args.class_only)
    print(f"Parsing completato in {time.time() - t0:.1f}s — {len(records)} record validi")

    if args.dry_run:
        print("\n--dry-run: nessuna scrittura su DB")
        print(stats.summary())
        return 0

    print(f"\nAggiornamento {len(records)} confezioni su Supabase...")
    t1 = time.time()
    updated = update_packages(records, batch_size=args.batch_size)
    print(f"  Update completato in {time.time() - t1:.1f}s — {updated} righe aggiornate")

    print(stats.summary())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
